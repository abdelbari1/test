from openpyxl import load_workbook, Workbook
from fashion.cache.mem_cache import MemCache
from fashion.models.item import Item
from fashion.models.user import User
from fashion.models.sizes import Sizes as SZ
import datetime
import shortuuid as suid
from fashion.models.domain import *

class XLDataLoader:
    
    def __init__(self) -> None:
        pass

    def load_default_model(self, cache: MemCache, fn: str):
        self.__load_all(cache, fn)

    def __load_all(self, cache: MemCache, fn: str):
        wb = load_workbook(fn)
        print('\tReading data...\n', end='')
        print('\tReading users...', end='')
        self.__loading_users(wb, cache, 'users')
        print('done. \n\tReading items...', end='')
        self.__loading_items(wb, cache, 'items')
        print('\tReading sizes...', end='')
        self.__loading_sizes(wb, cache, 'sizes')
        print('done')


    def __loading_items(self, wb: Workbook, cache: MemCache, shname: str):
        sheet = wb.get_sheet_by_name(shname)
        if sheet is not None and sheet.max_row > 1:
            its = [Item(
                    uid=sheet.cell(row=x, column=1).value, item_name=sheet.cell(row=x, column=2).value,
                    item_category=ItemCategory._value2member_map_.get(sheet.cell(row=x, column=3).value),
                    gender=Gender._value2member_map_.get(sheet.cell(row=x, column=4).value),
                    flash_sale=sheet.cell(row=x, column=5).value if sheet.cell(row=x, column=5).value is not None else 0,
                    actual_price=sheet.cell(row=x, column=6).value, currency=Currency._value2member_map_.get(sheet.cell(row=x, column=7).value), 
                    reference=sheet.cell(row=x, column=9).value if sheet.cell(row=x, column=9).value is not None else suid.random(length=6),
                    status_item=StatusItem._value2member_map_.get(sheet.cell(row=x, column=10).value), description=sheet.cell(row=x, column=11).value,
                    item_created=f'{datetime.datetime.now()}', item_model=sheet.cell(row=x, column=8).value,
                    user=cache.get_entity_by_id(sheet.cell(row=x, column=12).value, User)) 
                for x in range(2, sheet.max_row + 1)]
            cache.save_entities(its)

    def __loading_sizes(self, wb: Workbook, cache: MemCache, shname: str):
        sheet = wb.get_sheet_by_name(shname)
        items = set()
        if sheet is not None and sheet.max_row > 1:
            for x in range(2, sheet.max_row + 1):
                siz = SZ(uid=sheet.cell(row=x, column=1).value, size=sheet.cell(row=x, column=2).value,
                         qty=sheet.cell(row=x, column=3).value)
                item: Item = cache.get_entity_by_id(sheet.cell(row=x, column=4).value, Item)
                item.add_size(siz)
                items.add(item)
            for it in items:
                cache.save_sub_entity(it, Item.sizes, SZ, True)

    def __loading_users(self, wb: Workbook, cache: MemCache, shname: str):
        sheet = wb.get_sheet_by_name(shname)
        if sheet is not None and sheet.max_row > 1:
            usrs = [User(
                    uid=sheet.cell(row=u, column=1).value, fname=sheet.cell(row=u, column=2).value, lname=sheet.cell(row=u, column=3).value,
                    email=sheet.cell(row=u, column=4).value, password=sheet.cell(row=u, column=5).value, conPass=sheet.cell(row=u, column=6).value,
                    role=UserRole._value2member_map_.get(sheet.cell(row=u, column=7).value)) 
                for u in range(2, sheet.max_row + 1)]
            print(usrs)
            cache.save_entities(usrs)


    def __loading_wishlists():
        pass

    def __loading_notifications():
        pass